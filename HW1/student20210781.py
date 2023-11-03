#!/usr/bin/python3
from openpyxl import load_workbook

wb = load_workbook('student.xlsx')
ws = wb.active

total = ws.max_row - 1  # 헤더 행 제외
a = 0
a_plus = 0
b_ = 0
b_plus = 0
c = 0

for row in ws.iter_rows(min_row=2, values_only=True): # 첫번째 행은 헤더
    total = row[6]  
    if total < 40:
        grade = 'F'
    else:
        percentage = (total / 100) * 100  # 점수를 백분율로 변환

        if percentage >= 90 and a < total * 0.3:
            grade = 'A'
            a += 1
        elif percentage >= 85 and a_plus < total * 0.5 and a < total * 0.3:
            grade = 'A+'
            a_plus += 1
        elif percentage >= 70 and b < total * 0.7:
            grade = 'B'
            b += 1
        elif percentage >= 65 and b_plus < total * 0.5 and b < total * 0.7:
            grade = 'B+'
            b_plus += 1
        elif percentage >= 50 and c < total:
            grade = 'C'
            c += 1
        elif percentage >= 45 and c < total and c < total * 0.5:
            grade = 'C+'
            c += 1
        else:
            grade = 'F'

    # 변경된 학점 열에 학점 데이터 삽입
    ws.cell(row=row[0], column=8, value=grade)

# 변경된 데이터를 원래 엑셀 파일에 저장
wb.save('student.xlsx')
